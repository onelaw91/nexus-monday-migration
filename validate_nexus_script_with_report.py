import csv
import json
import os
import sys
import time

import requests


# ============================================================
# NEXUS CONSULTING GROUP - VALIDATION SCRIPT WITH REPORT OUTPUT
# ============================================================
#
# CODING LANGUAGE:
# This script is written in Python.
#
# IMPORTANT NOTE ABOUT CSV COLOR:
# This script creates:
#   1. nexus_validation_report.csv
#      - Simple plain-text report
#   2. nexus_validation_report.xlsx
#      - Formatted Excel report with black divider rows
#
#
# WHAT THIS SCRIPT VALIDATES:
# - Source engagement count vs monday engagement count
# - Source deliverable count vs monday deliverable count
# - Engagement IDs exist in monday
# - Deliverable IDs exist in monday
# - Deliverable required fields are populated
# - Deliverables are linked to engagements
# - Engagement statuses were normalized correctly
# - Deliverable statuses were normalized correctly
# - Budget and hours match the source CSV
#
# HOW TO RUN:
#   export MONDAY_API_TOKEN="your_api_token_here"
#   python3 validate_nexus_report.py
#
# OUTPUT FILES:
#   nexus_validation_report.csv
#   nexus_validation_report.xlsx
# ============================================================


# =========================
# CONFIGURATION
# =========================
#
# This section contains the values the script needs before it can run.
# The API token is intentionally NOT hardcoded in the script.
# It is read from your terminal environment variable for security.

MONDAY_API_URL = "https://api.monday.com/v2"
MONDAY_API_TOKEN = os.getenv("MONDAY_API_TOKEN")
MONDAY_API_VERSION = os.getenv("MONDAY_API_VERSION", "2025-10")

SOURCE_CSV = "nexus_smartsheet_export.csv"

CSV_REPORT_FILE = "nexus_validation_report.csv"
EXCEL_REPORT_FILE = "nexus_validation_report.xlsx"

# These should be the same board IDs used in your migration script.
ENGAGEMENTS_BOARD_ID = "18411674428"
DELIVERABLES_BOARD_ID = "18411677653"


# These are the monday.com column IDs from the Nexus Engagements board.
ENGAGEMENT_COLUMNS = {
    "engagement_id": "text_mm32y44j",
    "client": "text_mm32f83a",
    "engagement_lead": "text_mm32gqne",
    "start_date": "date",
    "end_date": "date_mm32kasr",
    "budget": "numeric_mm32q776",
    "status": "color_mm32y5vm",
}


# These are the monday.com column IDs from the Nexus Deliverables board.
DELIVERABLE_COLUMNS = {
    "deliverable_id": "text_mm322s08",
    "assignee": "text_mm32nd0j",
    "due_date": "date4",
    "priority": "color_mm32xg2s",
    "status": "color_mm32aqx",
    "hours_estimated": "numeric_mm32335k",
    "engagement_link": "board_relation_mm32d8pz",
}


# =========================
# STATUS NORMALIZATION
# =========================
#
# These dictionaries show how messy source values should be cleaned up.
# This should match the same logic used in the migration script.

ENGAGEMENT_STATUS_MAP = {
    "Active": "Active",
    "In Progress": "Active",
    "Complete": "Complete",
    "Done": "Complete",
    "On Hold": "On Hold",
    "Not Started": "Not Started",
}


DELIVERABLE_STATUS_MAP = {
    "To Do": "To Do",
    "Not Started": "To Do",
    "In Progress": "In Progress",
    "Working on it": "In Progress",
    "In Review": "In Review",
    "Done": "Done",
}


# =========================
# API AND FILE HELPERS
# =========================
#
# These functions do smaller utility jobs so the validation logic is easier
# to read and explain during the demo.


def require_env_vars():
    """
    Store the API token as an environment variable instead of hardcoding
    it into the script. This keeps credentials out of source code."
    """
    if not MONDAY_API_TOKEN:
        print("ERROR: Missing MONDAY_API_TOKEN environment variable.")
        print('Run this first: export MONDAY_API_TOKEN="your_api_token_here"')
        sys.exit(1)


def monday_request(query, variables=None):
    """
    Sends a GraphQL request to monday.com:

    This function is the central API helper. All monday.com reads go through
    here and it also handles complexity-limit retry behavior
    """
    headers = {
        "Authorization": MONDAY_API_TOKEN,
        "API-Version": MONDAY_API_VERSION,
        "Content-Type": "application/json",
    }

    payload = {
        "query": query,
        "variables": variables or {},
    }

    max_attempts = 5

    for attempt in range(max_attempts):
        response = requests.post(
            MONDAY_API_URL,
            headers=headers,
            json=payload,
            timeout=30,
        )

        try:
            data = response.json()
        except ValueError:
            raise Exception(f"Non-JSON response from monday.com: {response.text}")

        if "errors" in data:
            first_error = data["errors"][0]
            error_code = first_error.get("extensions", {}).get("code")

            if error_code == "COMPLEXITY_BUDGET_EXHAUSTED":
                retry_seconds = first_error.get("extensions", {}).get("retry_in_seconds", 20)
                print(f"Complexity limit hit. Waiting {retry_seconds} seconds before retrying...")
                time.sleep(retry_seconds + 2)
                continue

            raise Exception(json.dumps(data, indent=2))

        if response.status_code != 200:
            raise Exception(json.dumps(data, indent=2))

        return data["data"]

    raise Exception("monday.com API retry limit reached.")


def load_csv(path):
    """
    Reads the source CSV exported from Nexus Exported data 

    This is the source of truth. The validation script compares this file
    against what now exists in monday.com
    """
    with open(path, newline="", encoding="utf-8-sig") as file:
        return list(csv.DictReader(file))


def get_column_text(item, column_id):
    """
    Gets the visible text value from a monday.com column.

    This works for normal fields like text, status, dates, and numbers.
    Connect-board relationships are validated separately.
    """
    for column in item["column_values"]:
        if column["id"] == column_id:
            return column.get("text") or ""
    return ""


def normalize_money(value):
    """
    Normalizes money values so formatting differences do not cause false errors.

    Example:
    "150000", "150,000", and "$150000" should compare as the same value.
    """
    if value in (None, ""):
        return ""

    return str(int(float(str(value).replace(",", "").replace("$", ""))))


def normalize_number_text(value):
    """
    Normalizes numeric values like estimated hours.

    Example:
    "40", "40.0", and "40.00" should compare as the same value.
    """
    if value in (None, ""):
        return ""

    try:
        return str(int(float(str(value).replace(",", ""))))
    except ValueError:
        return str(value)


def add_report_row(report_rows, section, validation_check, record, source_data,
                   monday_data, result, details):
    """
    Adds one row to the validation report.

    Simplified report columns:
    - Section
    - Validation Check
    - Record
    - Source Data
    - monday.com Data
    - Result
    - Details
    """
    report_rows.append({
        "Section": section,
        "Validation Check": validation_check,
        "Record": record,
        "Source Data": source_data,
        "monday.com Data": monday_data,
        "Result": result,
        "Details": details,
    })


# =========================
# MONDAY READ FUNCTIONS
# =========================
#
# These functions pull migrated data back out of monday.com.


def get_board_items(board_id):
    """
    Gets all items from a monday.com board using the GraphQL API.

    Independently query monday.com to confirm what actually loaded.
    "
    """
    query = """
    query GetBoardItems($board_id: ID!, $cursor: String) {
      boards(ids: [$board_id]) {
        items_page(limit: 500, cursor: $cursor) {
          cursor
          items {
            id
            name
            column_values {
              id
              text
            }
          }
        }
      }
    }
    """

    all_items = []
    cursor = None

    while True:
        variables = {
            "board_id": str(board_id),
            "cursor": cursor,
        }

        data = monday_request(query, variables)

        if not data["boards"]:
            raise Exception(
                f"No board found for board ID {board_id}. "
                "Check that the board ID is correct."
            )

        page = data["boards"][0]["items_page"]

        all_items.extend(page["items"])
        cursor = page["cursor"]

        if not cursor:
            break

    return all_items


def get_linked_engagements_for_deliverables(deliverable_items):
    """
    Gets the engagement links for each deliverable.

    Important:
    Connect-board fields are relationship fields, not normal text fields.
    That is why this script uses linked_items instead of reading text.

    This is how I validate that each deliverable is actually connected to
    its parent engagement."
    """
    if not deliverable_items:
        return {}

    item_ids = [str(item["id"]) for item in deliverable_items]

    query = """
    query GetLinkedEngagements(
      $item_ids: [ID!],
      $linked_board_id: ID!,
      $column_id: String!
    ) {
      items(ids: $item_ids) {
        id
        name
        linked_items(
          linked_board_id: $linked_board_id,
          link_to_item_column_id: $column_id
        ) {
          id
          name
        }
      }
    }
    """

    variables = {
        "item_ids": item_ids,
        "linked_board_id": str(ENGAGEMENTS_BOARD_ID),
        "column_id": DELIVERABLE_COLUMNS["engagement_link"],
    }

    data = monday_request(query, variables)

    result = {}

    for item in data["items"]:
        result[str(item["id"])] = item.get("linked_items", [])

    return result


# =========================
# VALIDATION FUNCTIONS
# =========================
#
# Each function validates one area of the migration and adds rows to the
# final report.


def validate_counts(source_rows, engagement_items, deliverable_items, errors, report_rows):
    section = "1. Record Counts"

    source_engagement_count = len(set(row["engagement_id"] for row in source_rows))
    monday_engagement_count = len(engagement_items)

    if source_engagement_count == monday_engagement_count:
        result = "PASS"
        details = "Source and monday.com engagement counts match."
    else:
        result = "FAIL"
        details = "Source and monday.com engagement counts do not match."
        errors.append(details)

    add_report_row(
        report_rows,
        section,
        "Engagement count matches",
        "All engagements",
        source_engagement_count,
        monday_engagement_count,
        result,
        details,
    )

    source_deliverable_count = len(set(row["deliverable_id"] for row in source_rows))
    monday_deliverable_count = len(deliverable_items)

    if source_deliverable_count == monday_deliverable_count:
        result = "PASS"
        details = "Source and monday.com deliverable counts match."
    else:
        result = "FAIL"
        details = "Source and monday.com deliverable counts do not match."
        errors.append(details)

    add_report_row(
        report_rows,
        section,
        "Deliverable count matches",
        "All deliverables",
        source_deliverable_count,
        monday_deliverable_count,
        result,
        details,
    )


def validate_engagement_ids(source_rows, engagement_items, errors, report_rows):
    section = "2. Engagement IDs"

    source_engagements = {}

    for row in source_rows:
        if row["engagement_id"] not in source_engagements:
            source_engagements[row["engagement_id"]] = row

    monday_ids = set(
        get_column_text(item, ENGAGEMENT_COLUMNS["engagement_id"])
        for item in engagement_items
    )

    for engagement_id, row in sorted(source_engagements.items()):
        if engagement_id in monday_ids:
            result = "PASS"
            monday_data = "Found"
            details = "Engagement ID exists in monday.com."
        else:
            result = "FAIL"
            monday_data = "Missing"
            details = "Engagement ID is missing in monday.com."
            errors.append(f"{engagement_id} is missing in monday.com")

        add_report_row(
            report_rows,
            section,
            "Engagement ID exists",
            f"{engagement_id} - {row['engagement_name']}",
            "Found in source CSV",
            monday_data,
            result,
            details,
        )


def validate_deliverable_ids(source_rows, deliverable_items, errors, report_rows):
    section = "3. Deliverable IDs"

    source_deliverables = {
        row["deliverable_id"]: row
        for row in source_rows
    }

    monday_ids = set(
        get_column_text(item, DELIVERABLE_COLUMNS["deliverable_id"])
        for item in deliverable_items
    )

    for deliverable_id, row in sorted(source_deliverables.items()):
        if deliverable_id in monday_ids:
            result = "PASS"
            monday_data = "Found"
            details = "Deliverable ID exists in monday.com."
        else:
            result = "FAIL"
            monday_data = "Missing"
            details = "Deliverable ID is missing in monday.com."
            errors.append(f"{deliverable_id} is missing in monday.com")

        add_report_row(
            report_rows,
            section,
            "Deliverable ID exists",
            f"{deliverable_id} - {row['deliverable_name']}",
            "Found in source CSV",
            monday_data,
            result,
            details,
        )


def validate_relationships(source_rows, deliverable_items, linked_engagements_by_item_id,
                           errors, report_rows):
    section = "4. Deliverable to Engagement Links"

    source_by_deliverable = {
        row["deliverable_id"]: row
        for row in source_rows
    }

    for item in deliverable_items:
        deliverable_id = get_column_text(item, DELIVERABLE_COLUMNS["deliverable_id"])
        source_row = source_by_deliverable.get(deliverable_id)

        if not source_row:
            continue

        expected_engagement = source_row["engagement_name"]
        linked_items = linked_engagements_by_item_id.get(str(item["id"]), [])
        linked_names = [linked_item["name"] for linked_item in linked_items]
        actual_engagements = ", ".join(linked_names) if linked_names else "No linked engagement"

        if expected_engagement in linked_names:
            result = "PASS"
            details = "Deliverable is linked to the expected engagement."
        else:
            result = "FAIL"
            details = "Deliverable is not linked to the expected engagement."
            errors.append(
                f"{deliverable_id} expected link '{expected_engagement}' but found '{actual_engagements}'"
            )

        add_report_row(
            report_rows,
            section,
            "Deliverable linked to correct engagement",
            f"{deliverable_id} - {item['name']}",
            expected_engagement,
            actual_engagements,
            result,
            details,
        )


def validate_required_fields(deliverable_items, warnings, report_rows):
    section = "5. Required Deliverable Fields"

    required_fields = [
        ("Assignee populated", "Assignee", DELIVERABLE_COLUMNS["assignee"]),
        ("Due date populated", "Due Date", DELIVERABLE_COLUMNS["due_date"]),
        ("Hours populated", "Hours Estimated", DELIVERABLE_COLUMNS["hours_estimated"]),
    ]

    for item in deliverable_items:
        deliverable_id = get_column_text(item, DELIVERABLE_COLUMNS["deliverable_id"]) or item["name"]

        for check_name, friendly_field_name, column_id in required_fields:
            actual_value = get_column_text(item, column_id)

            if actual_value:
                result = "PASS"
                details = f"{friendly_field_name} is populated."
                monday_data = actual_value
            else:
                result = "WARN"
                details = f"{friendly_field_name} is blank."
                monday_data = "Blank"
                warnings.append(f"{deliverable_id} is missing {friendly_field_name}")

            add_report_row(
                report_rows,
                section,
                check_name,
                f"{deliverable_id} - {item['name']}",
                "Should be populated",
                monday_data,
                result,
                details,
            )


def validate_engagement_statuses(source_rows, engagement_items, errors, report_rows):
    section = "6. Engagement Status Normalization"

    source_engagements = {}

    for row in source_rows:
        if row["engagement_id"] not in source_engagements:
            source_engagements[row["engagement_id"]] = row

    monday_by_id = {
        get_column_text(item, ENGAGEMENT_COLUMNS["engagement_id"]): item
        for item in engagement_items
    }

    for engagement_id, source_row in sorted(source_engagements.items()):
        expected_status = ENGAGEMENT_STATUS_MAP[source_row["engagement_status"]]
        monday_item = monday_by_id.get(engagement_id)

        if not monday_item:
            continue

        actual_status = get_column_text(monday_item, ENGAGEMENT_COLUMNS["status"])

        if actual_status == expected_status:
            result = "PASS"
            details = "Engagement status was normalized correctly."
        else:
            result = "FAIL"
            details = "Engagement status does not match expected normalized value."
            errors.append(
                f"{engagement_id} expected status '{expected_status}' but found '{actual_status}'"
            )

        add_report_row(
            report_rows,
            section,
            "Engagement status matches normalized value",
            f"{engagement_id} - {monday_item['name']}",
            f"{source_row['engagement_status']} -> {expected_status}",
            actual_status,
            result,
            details,
        )


def validate_deliverable_statuses(source_rows, deliverable_items, errors, report_rows):
    section = "7. Deliverable Status Normalization"

    source_deliverables = {
        row["deliverable_id"]: row
        for row in source_rows
    }

    monday_by_id = {
        get_column_text(item, DELIVERABLE_COLUMNS["deliverable_id"]): item
        for item in deliverable_items
    }

    for deliverable_id, source_row in sorted(source_deliverables.items()):
        expected_status = DELIVERABLE_STATUS_MAP[source_row["deliverable_status"]]
        monday_item = monday_by_id.get(deliverable_id)

        if not monday_item:
            continue

        actual_status = get_column_text(monday_item, DELIVERABLE_COLUMNS["status"])

        if actual_status == expected_status:
            result = "PASS"
            details = "Deliverable status was normalized correctly."
        else:
            result = "FAIL"
            details = "Deliverable status does not match expected normalized value."
            errors.append(
                f"{deliverable_id} expected status '{expected_status}' but found '{actual_status}'"
            )

        add_report_row(
            report_rows,
            section,
            "Deliverable status matches normalized value",
            f"{deliverable_id} - {monday_item['name']}",
            f"{source_row['deliverable_status']} -> {expected_status}",
            actual_status,
            result,
            details,
        )


def validate_budget_and_hours(source_rows, engagement_items, deliverable_items,
                              errors, report_rows):
    section = "8. Budget and Hours"

    source_engagements = {}

    for row in source_rows:
        if row["engagement_id"] not in source_engagements:
            source_engagements[row["engagement_id"]] = row

    monday_engagements = {
        get_column_text(item, ENGAGEMENT_COLUMNS["engagement_id"]): item
        for item in engagement_items
    }

    for engagement_id, source_row in sorted(source_engagements.items()):
        monday_item = monday_engagements.get(engagement_id)

        if not monday_item:
            continue

        expected_budget = normalize_money(source_row["budget"])
        actual_budget = normalize_number_text(
            get_column_text(monday_item, ENGAGEMENT_COLUMNS["budget"])
        )

        if expected_budget == actual_budget:
            result = "PASS"
            details = "Budget matches the source CSV."
        else:
            result = "FAIL"
            details = "Budget does not match the source CSV."
            errors.append(
                f"{engagement_id} expected budget '{expected_budget}' but found '{actual_budget}'"
            )

        add_report_row(
            report_rows,
            section,
            "Budget matches source",
            f"{engagement_id} - {monday_item['name']}",
            expected_budget,
            actual_budget,
            result,
            details,
        )

    source_deliverables = {
        row["deliverable_id"]: row
        for row in source_rows
    }

    monday_deliverables = {
        get_column_text(item, DELIVERABLE_COLUMNS["deliverable_id"]): item
        for item in deliverable_items
    }

    for deliverable_id, source_row in sorted(source_deliverables.items()):
        monday_item = monday_deliverables.get(deliverable_id)

        if not monday_item:
            continue

        expected_hours = normalize_number_text(source_row["hours_estimated"])
        actual_hours = normalize_number_text(
            get_column_text(monday_item, DELIVERABLE_COLUMNS["hours_estimated"])
        )

        if expected_hours == actual_hours:
            result = "PASS"
            details = "Estimated hours match the source CSV."
        else:
            result = "FAIL"
            details = "Estimated hours do not match the source CSV."
            errors.append(
                f"{deliverable_id} expected hours '{expected_hours}' but found '{actual_hours}'"
            )

        add_report_row(
            report_rows,
            section,
            "Estimated hours match source",
            f"{deliverable_id} - {monday_item['name']}",
            expected_hours,
            actual_hours,
            result,
            details,
        )


# =========================
# REPORT WRITING FUNCTIONS
# =========================


def write_csv_report(report_rows):
    """
    Writes the simple CSV report.

    CSV cannot store colors, so the black divider rows only appear in the
    Excel version.
    """
    fieldnames = [
        "Section",
        "Validation Check",
        "Record",
        "Source Data",
        "monday.com Data",
        "Result",
        "Details",
    ]

    with open(CSV_REPORT_FILE, "w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report_rows)


def write_excel_report(report_rows):
    """
    Writes a formatted Excel report.

    This creates black divider rows between each validation category and
    color-codes PASS, WARN, and FAIL results.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Excel report skipped because openpyxl is not installed.")
        print("To enable Excel output, run: python3 -m pip install openpyxl")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Validation Report"

    headers = [
        "Validation Check",
        "Record",
        "Source Data",
        "monday.com Data",
        "Result",
        "Details",
    ]

    black_fill = PatternFill("solid", fgColor="000000")
    white_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    pass_fill = PatternFill("solid", fgColor="D9EAD3")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    current_row = 1
    current_section = None

    for row in report_rows:
        section = row["Section"]

        if section != current_section:
            current_section = section

            sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(headers),
            )
            section_cell = sheet.cell(row=current_row, column=1)
            section_cell.value = section
            section_cell.fill = black_fill
            section_cell.font = white_font
            section_cell.alignment = Alignment(horizontal="left")
            current_row += 1

            for col_index, header in enumerate(headers, start=1):
                cell = sheet.cell(row=current_row, column=col_index)
                cell.value = header
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            current_row += 1

        values = [
            row["Validation Check"],
            row["Record"],
            row["Source Data"],
            row["monday.com Data"],
            row["Result"],
            row["Details"],
        ]

        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=current_row, column=col_index)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if headers[col_index - 1] == "Result":
                if value == "PASS":
                    cell.fill = pass_fill
                elif value == "WARN":
                    cell.fill = warn_fill
                elif value == "FAIL":
                    cell.fill = fail_fill

        current_row += 1

    column_widths = {
        "A": 34,
        "B": 42,
        "C": 28,
        "D": 28,
        "E": 12,
        "F": 52,
    }

    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(EXCEL_REPORT_FILE)


def print_terminal_summary(source_rows, engagement_items, deliverable_items,
                           linked_engagements_by_item_id, errors, warnings):
    """
    Prints the high-level result to the terminal.
    """
    source_engagement_count = len(set(row["engagement_id"] for row in source_rows))
    source_deliverable_count = len(set(row["deliverable_id"] for row in source_rows))

    linked_count = 0
    for item in deliverable_items:
        if linked_engagements_by_item_id.get(str(item["id"])):
            linked_count += 1

    print("\nNEXUS MIGRATION VALIDATION SUMMARY")
    print("=" * 50)
    print(f"Source engagements:             {source_engagement_count}")
    print(f"monday.com engagements:         {len(engagement_items)}")
    print(f"Source deliverables:            {source_deliverable_count}")
    print(f"monday.com deliverables:        {len(deliverable_items)}")
    print(f"Deliverables linked correctly:  {linked_count} / {len(deliverable_items)}")

    print("\nReport Files")
    print("-" * 50)
    print(f"CSV report:   {CSV_REPORT_FILE}")
    print(f"Excel report: {EXCEL_REPORT_FILE}")

    print("\nValidation Result")
    print("-" * 50)

    if errors:
        print("FAILED")
        for error in errors:
            print(f"ERROR: {error}")
    else:
        print("PASSED")

    if warnings:
        print("\nWarnings")
        for warning in warnings:
            print(f"WARNING: {warning}")

    print("=" * 50)


# =========================
# MAIN WORKFLOW
# =========================
#
# This is the main storyline of the script.

def main():
    require_env_vars()

    print("Starting Nexus validation...")

    # Step 1: Read the original CSV.
    source_rows = load_csv(SOURCE_CSV)
    print(f"Loaded {len(source_rows)} source rows from {SOURCE_CSV}")

    # Step 2: Read migrated records from monday.com.
    engagement_items = get_board_items(ENGAGEMENTS_BOARD_ID)
    deliverable_items = get_board_items(DELIVERABLES_BOARD_ID)
    print(f"Retrieved {len(engagement_items)} engagement items from monday.com")
    print(f"Retrieved {len(deliverable_items)} deliverable items from monday.com")

    # Step 3: Read deliverable-to-engagement relationships from monday.com.
    linked_engagements_by_item_id = get_linked_engagements_for_deliverables(deliverable_items)

    # Step 4: Store validation results.
    errors = []
    warnings = []
    report_rows = []

    # Step 5: Run validation checks.
    validate_counts(source_rows, engagement_items, deliverable_items, errors, report_rows)
    validate_engagement_ids(source_rows, engagement_items, errors, report_rows)
    validate_deliverable_ids(source_rows, deliverable_items, errors, report_rows)
    validate_relationships(
        source_rows,
        deliverable_items,
        linked_engagements_by_item_id,
        errors,
        report_rows,
    )
    validate_required_fields(deliverable_items, warnings, report_rows)
    validate_engagement_statuses(source_rows, engagement_items, errors, report_rows)
    validate_deliverable_statuses(source_rows, deliverable_items, errors, report_rows)
    validate_budget_and_hours(source_rows, engagement_items, deliverable_items, errors, report_rows)

    # Step 6: Write report files.
    write_csv_report(report_rows)
    write_excel_report(report_rows)

    # Step 7: Print high-level result.
    print_terminal_summary(
        source_rows,
        engagement_items,
        deliverable_items,
        linked_engagements_by_item_id,
        errors,
        warnings,
    )


if __name__ == "__main__":
    main()
